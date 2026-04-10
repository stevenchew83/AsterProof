# ruff: noqa: INP001

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from dataclasses import field
from datetime import date
from datetime import datetime
from io import BytesIO
from io import StringIO
from pathlib import Path
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from inspinia.rankings.models import ImportBatch
from inspinia.rankings.models import ImportRowIssue
from inspinia.rankings.models import Student
from inspinia.rankings.models import normalize_name
from inspinia.rankings.models import normalize_whitespace
from inspinia.users.roles import user_has_admin_role

CSV_ISSUE_MESSAGE = "Unsupported upload format."
AMBIGUOUS_EXTERNAL_CODE = "AMBIGUOUS_EXTERNAL_CODE"
AMBIGUOUS_FULL_NRIC = "AMBIGUOUS_FULL_NRIC"
AMBIGUOUS_NAME_BIRTH_YEAR = "AMBIGUOUS_NAME_BIRTH_YEAR"
INVALID_BIRTH_YEAR = "INVALID_BIRTH_YEAR"
INVALID_ACTIVE_VALUE = "INVALID_ACTIVE_VALUE"
MISSING_FULL_NAME = "MISSING_FULL_NAME"
FILE_EXTENSIONS = {".csv", ".xlsx"}

HEADER_ALIASES = {
    "active": "active",
    "birth_year": "birth_year",
    "date_of_birth": "date_of_birth",
    "dob": "date_of_birth",
    "external_code": "external_code",
    "full_name": "full_name",
    "full_nric": "full_nric",
    "gender": "gender",
    "legacy_code": "legacy_code",
    "masked_nric": "masked_nric",
    "name": "full_name",
    "notes": "notes",
    "state": "state",
    "student_code": "external_code",
    "student_name": "full_name",
}

TRUE_TOKENS = {"1", "true", "yes", "y", "on", "active"}
FALSE_TOKENS = {"0", "false", "no", "n", "off", "inactive"}


@dataclass(slots=True)
class StudentMasterRowData:
    full_name: str
    birth_year: int | None
    date_of_birth: date | None
    gender: str
    state: str
    masked_nric: str
    full_nric: str
    external_code: str
    legacy_code: str
    active: bool | None
    notes: str
    normalized_name: str

    @classmethod
    def from_mapping(cls, mapping: dict[str, Any]) -> tuple[StudentMasterRowData, list[dict[str, str]]]:
        issues: list[dict[str, str]] = []

        full_name = normalize_whitespace(mapping.get("full_name", ""))
        if not full_name:
            issues.append(
                {
                    "issue_code": MISSING_FULL_NAME,
                    "message": "Full name is required.",
                },
            )

        birth_year, birth_year_issue = _parse_birth_year(mapping.get("birth_year"))
        if birth_year_issue is not None:
            issues.append(birth_year_issue)

        date_of_birth, dob_issue = _parse_date_of_birth(mapping.get("date_of_birth"))
        if dob_issue is not None:
            issues.append(dob_issue)

        active, active_issue = _parse_active_flag(mapping.get("active"))
        if active_issue is not None:
            issues.append(active_issue)

        return (
            cls(
                full_name=full_name,
                birth_year=birth_year,
                date_of_birth=date_of_birth,
                gender=normalize_whitespace(mapping.get("gender", "")),
                state=normalize_whitespace(mapping.get("state", "")),
                masked_nric=normalize_whitespace(mapping.get("masked_nric", "")),
                full_nric=normalize_whitespace(mapping.get("full_nric", "")),
                external_code=normalize_whitespace(mapping.get("external_code", "")).upper(),
                legacy_code=normalize_whitespace(mapping.get("legacy_code", "")),
                active=active,
                notes=normalize_whitespace(mapping.get("notes", "")),
                normalized_name=normalize_name(full_name),
            ),
            issues,
        )


@dataclass(slots=True)
class StudentMasterPreviewRow:
    row_number: int
    raw_row_json: dict[str, Any]
    data: StudentMasterRowData
    bucket: str
    match_strategy: str = ""
    matched_student_id: int | None = None
    issue_code: str = ""
    issue_message: str = ""


@dataclass(slots=True)
class StudentMasterImportPreview:
    import_batch_id: int
    rows: list[StudentMasterPreviewRow] = field(default_factory=list)
    match_count: int = 0
    create_count: int = 0
    ambiguous_count: int = 0
    error_count: int = 0
    matched_by_external_code: int = 0
    matched_by_full_nric: int = 0
    matched_by_name_birth_year: int = 0

    @property
    def rows_processed(self) -> int:
        return len(self.rows)

    @property
    def skipped_count(self) -> int:
        return self.ambiguous_count + self.error_count

    def to_summary_json(self) -> dict[str, Any]:
        return {
            "rows_processed": self.rows_processed,
            "match_count": self.match_count,
            "create_count": self.create_count,
            "ambiguous_count": self.ambiguous_count,
            "error_count": self.error_count,
            "matched_by_external_code": self.matched_by_external_code,
            "matched_by_full_nric": self.matched_by_full_nric,
            "matched_by_name_birth_year": self.matched_by_name_birth_year,
        }


@dataclass(slots=True)
class StudentMasterImportApplyResult:
    created: int
    updated: int
    skipped: int

    def to_summary_json(self, preview: StudentMasterImportPreview) -> dict[str, Any]:
        return {
            "rows_processed": preview.rows_processed,
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "match_count": preview.match_count,
            "create_count": preview.create_count,
            "ambiguous_count": preview.ambiguous_count,
            "error_count": preview.error_count,
            "matched_by_external_code": preview.matched_by_external_code,
            "matched_by_full_nric": preview.matched_by_full_nric,
            "matched_by_name_birth_year": preview.matched_by_name_birth_year,
        }


@dataclass(slots=True)
class StudentMatchResult:
    bucket: str
    strategy: str = ""
    student: Student | None = None
    issue_code: str = ""
    issue_message: str = ""


def preview_student_master_import(*, import_batch: ImportBatch, actor: object | None) -> StudentMasterImportPreview:
    raw_rows = list(_load_import_rows(import_batch.uploaded_file))
    issue_rows: list[ImportRowIssue] = []
    preview_rows: list[StudentMasterPreviewRow] = []
    match_count = 0
    create_count = 0
    ambiguous_count = 0
    error_count = 0
    matched_by_external_code = 0
    matched_by_full_nric = 0
    matched_by_name_birth_year = 0

    import_batch.row_issues.all().delete()

    for row_number, raw_row in raw_rows:
        data, parse_issues = StudentMasterRowData.from_mapping(raw_row)
        if parse_issues:
            issue = parse_issues[0]
            error_count += 1
            preview_rows.append(
                StudentMasterPreviewRow(
                    row_number=row_number,
                    raw_row_json=raw_row,
                    data=data,
                    bucket="error",
                    issue_code=issue["issue_code"],
                    issue_message=issue["message"],
                ),
            )
            issue_rows.extend(
                _build_issue_rows(
                    import_batch=import_batch,
                    row_number=row_number,
                    severity=ImportRowIssue.Severity.ERROR,
                    raw_row_json=raw_row,
                    issues=parse_issues,
                ),
            )
            continue

        match_result = _match_student_row(data, actor=actor)
        if match_result.bucket == "matched" and match_result.student is not None:
            match_count += 1
            if match_result.strategy == "external_code":
                matched_by_external_code += 1
            elif match_result.strategy == "full_nric":
                matched_by_full_nric += 1
            elif match_result.strategy == "normalized_name_birth_year":
                matched_by_name_birth_year += 1

            preview_rows.append(
                StudentMasterPreviewRow(
                    row_number=row_number,
                    raw_row_json=raw_row,
                    data=data,
                    bucket="matched",
                    match_strategy=match_result.strategy,
                    matched_student_id=match_result.student.id,
                ),
            )
            continue

        if match_result.bucket == "ambiguous":
            ambiguous_count += 1
            preview_rows.append(
                StudentMasterPreviewRow(
                    row_number=row_number,
                    raw_row_json=raw_row,
                    data=data,
                    bucket="ambiguous",
                    match_strategy=match_result.strategy,
                    issue_code=match_result.issue_code,
                    issue_message=match_result.issue_message,
                ),
            )
            issue_rows.extend(
                _build_issue_rows(
                    import_batch=import_batch,
                    row_number=row_number,
                    severity=ImportRowIssue.Severity.WARNING,
                    raw_row_json=raw_row,
                    issues=[
                        {
                            "issue_code": match_result.issue_code,
                            "message": match_result.issue_message,
                        },
                    ],
                ),
            )
            continue

        create_count += 1
        preview_rows.append(
            StudentMasterPreviewRow(
                row_number=row_number,
                raw_row_json=raw_row,
                data=data,
                bucket="create",
            ),
        )

    if issue_rows:
        ImportRowIssue.objects.bulk_create(issue_rows)

    preview = StudentMasterImportPreview(
        import_batch_id=import_batch.id,
        rows=preview_rows,
        match_count=match_count,
        create_count=create_count,
        ambiguous_count=ambiguous_count,
        error_count=error_count,
        matched_by_external_code=matched_by_external_code,
        matched_by_full_nric=matched_by_full_nric,
        matched_by_name_birth_year=matched_by_name_birth_year,
    )
    import_batch.status = ImportBatch.Status.PREVIEWED
    import_batch.summary_json = preview.to_summary_json()
    import_batch.save(update_fields=["status", "summary_json"])
    return preview


def apply_student_master_import(
    *,
    preview: StudentMasterImportPreview,
    import_batch: ImportBatch,
    actor: object | None,
) -> StudentMasterImportApplyResult:
    if preview.import_batch_id != import_batch.id:
        msg = "Preview and batch must refer to the same import batch."
        raise ValueError(msg)

    created = 0
    updated = 0

    with transaction.atomic():
        for row in preview.rows:
            if row.bucket == "matched":
                if row.matched_student_id is None:
                    msg = "Matched preview rows must carry a student id."
                    raise ValueError(msg)
                student = Student.objects.select_for_update().get(pk=row.matched_student_id)
                _apply_row_to_student(student, row.data, actor=actor)
                student.save()
                updated += 1
                continue

            if row.bucket == "create":
                student = Student()
                _apply_row_to_student(student, row.data, actor=actor)
                student.save()
                created += 1

    result = StudentMasterImportApplyResult(
        created=created,
        updated=updated,
        skipped=preview.skipped_count,
    )
    import_batch.status = (
        ImportBatch.Status.PARTIAL
        if preview.skipped_count
        else ImportBatch.Status.APPLIED
    )
    import_batch.summary_json = result.to_summary_json(preview)
    import_batch.save(update_fields=["status", "summary_json"])
    return result


def _apply_row_to_student(student: Student, row: StudentMasterRowData, *, actor: object | None) -> None:
    student.full_name = row.full_name
    student.birth_year = row.birth_year if row.birth_year is not None else student.birth_year
    student.date_of_birth = row.date_of_birth if row.date_of_birth is not None else student.date_of_birth
    student.active = row.active if row.active is not None else student.active

    for field_name, value in (
        ("gender", row.gender),
        ("state", row.state),
        ("masked_nric", row.masked_nric),
        ("external_code", row.external_code),
        ("legacy_code", row.legacy_code),
        ("notes", row.notes),
    ):
        if value:
            setattr(student, field_name, value)

    if _can_use_full_nric(actor) and row.full_nric:
        student.full_nric = row.full_nric


def _match_student_row(row: StudentMasterRowData, *, actor: object | None) -> StudentMatchResult:
    if row.external_code:
        result = _match_unique_students(
            list(Student.objects.filter(external_code=row.external_code)),
            strategy="external_code",
            issue_code=AMBIGUOUS_EXTERNAL_CODE,
            issue_message=f"Multiple students match external code {row.external_code}.",
        )
        if result is not None:
            return result

    if _can_use_full_nric(actor) and row.full_nric:
        result = _match_unique_students(
            list(Student.objects.filter(full_nric=row.full_nric)),
            strategy="full_nric",
            issue_code=AMBIGUOUS_FULL_NRIC,
            issue_message=f"Multiple students match full NRIC {row.full_nric}.",
        )
        if result is not None:
            return result

    if row.normalized_name and row.birth_year is not None:
        result = _match_unique_students(
            list(
                Student.objects.filter(
                    normalized_name=row.normalized_name,
                    birth_year=row.birth_year,
                ),
            ),
            strategy="normalized_name_birth_year",
            issue_code=AMBIGUOUS_NAME_BIRTH_YEAR,
            issue_message="Multiple students match normalized name + birth year.",
        )
        if result is not None:
            return result

    return StudentMatchResult(bucket="create")


def _load_import_rows(upload: Any) -> list[tuple[int, dict[str, Any]]]:
    filename = normalize_whitespace(getattr(upload, "name", ""))
    if Path(filename).suffix.lower() not in FILE_EXTENSIONS:
        raise ValueError(CSV_ISSUE_MESSAGE)

    payload = _read_upload_bytes(upload)
    if Path(filename).suffix.lower() == ".csv":
        return _load_csv_rows(payload)
    return _load_xlsx_rows(payload)


def _load_csv_rows(payload: bytes) -> list[tuple[int, dict[str, Any]]]:
    text = payload.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in enumerate(reader, start=2):
        normalized = _normalize_row_mapping(row)
        if _row_has_data(normalized):
            rows.append((row_number, normalized))
    return rows


def _load_xlsx_rows(payload: bytes) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[tuple[int, dict[str, Any]]] = []
    headers: list[str] = []

    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number == 1:
            headers = [normalize_whitespace(value) for value in values]
            continue
        row = dict(zip(headers, values, strict=False))
        normalized = _normalize_row_mapping(row)
        if _row_has_data(normalized):
            rows.append((row_number, normalized))

    workbook.close()
    return rows


def _normalize_row_mapping(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        canonical_key = HEADER_ALIASES.get(_canonical_header(key))
        if canonical_key is None:
            continue
        normalized[canonical_key] = _normalize_cell_value(value)
    return normalized


def _canonical_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", normalize_whitespace(str(value)).casefold()).strip("_")


def _normalize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return normalize_whitespace(value)
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _row_has_data(row: dict[str, Any]) -> bool:
    return any(normalize_whitespace(str(value)) for value in row.values())


def _parse_birth_year(value: Any) -> tuple[int | None, dict[str, str] | None]:
    issue = _invalid_birth_year_issue()
    parsed: int | None = None

    if value in {None, ""}:
        pass
    elif isinstance(value, bool):
        return None, issue
    elif isinstance(value, int):
        if value > 0:
            parsed = value
        else:
            return None, issue
    elif isinstance(value, float):
        if value.is_integer() and value > 0:
            parsed = int(value)
        else:
            return None, issue
    else:
        text = normalize_whitespace(str(value))
        if text:
            try:
                parsed = int(text)
            except ValueError:
                return None, issue
            if parsed <= 0:
                return None, issue

    return parsed, None


def _parse_date_of_birth(value: Any) -> tuple[date | None, dict[str, str] | None]:
    if value in {None, ""}:
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    text = normalize_whitespace(str(value))
    if not text:
        return None, None
    try:
        return date.fromisoformat(text), None
    except ValueError:
        return None, {"issue_code": "INVALID_DATE_OF_BIRTH", "message": "Date of birth must be YYYY-MM-DD."}


def _parse_active_flag(value: Any) -> tuple[bool | None, dict[str, str] | None]:
    if value in {None, ""}:
        return None, None
    if isinstance(value, bool):
        return value, None

    token = normalize_whitespace(str(value)).casefold()
    if token in TRUE_TOKENS:
        return True, None
    if token in FALSE_TOKENS:
        return False, None
    return None, {"issue_code": INVALID_ACTIVE_VALUE, "message": "Active must be true or false."}


def _build_issue_rows(
    *,
    import_batch: ImportBatch,
    row_number: int,
    severity: str,
    raw_row_json: dict[str, Any],
    issues: list[dict[str, str]],
) -> list[ImportRowIssue]:
    return [
        ImportRowIssue(
            import_batch=import_batch,
            row_number=row_number,
            severity=severity,
            issue_code=issue["issue_code"],
            message=issue["message"],
            raw_row_json=raw_row_json,
        )
        for issue in issues
    ]


def _can_use_full_nric(actor: object | None) -> bool:
    return user_has_admin_role(actor)


def _read_upload_bytes(upload: Any) -> bytes:
    if hasattr(upload, "open"):
        upload.open("rb")
    try:
        if hasattr(upload, "seek"):
            upload.seek(0)
        payload = upload.read()
    finally:
        if hasattr(upload, "close"):
            from contextlib import suppress

            with suppress(Exception):
                upload.close()
    if not isinstance(payload, bytes):
        return bytes(payload, encoding="utf-8")
    return payload


def _invalid_birth_year_issue() -> dict[str, str]:
    return {"issue_code": INVALID_BIRTH_YEAR, "message": "Birth year must be an integer."}


def _match_unique_students(
    matches: list[Student],
    *,
    strategy: str,
    issue_code: str,
    issue_message: str,
) -> StudentMatchResult | None:
    if len(matches) == 1:
        return StudentMatchResult(bucket="matched", strategy=strategy, student=matches[0])
    if len(matches) > 1:
        return StudentMatchResult(
            bucket="ambiguous",
            strategy=strategy,
            issue_code=issue_code,
            issue_message=issue_message,
        )
    return None
